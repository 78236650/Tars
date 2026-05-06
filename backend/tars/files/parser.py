"""TARS Files - 文件解析器"""
import base64
import csv
import io
from pathlib import Path
from typing import Optional

from .models import FileRecord, ParsedContent
from .storage import TEXT_EXTENSIONS

MAX_TEXT_LENGTH = 30000
MAX_OCR_LENGTH = 5000


class FileParser:
    """根据文件类型解析内容"""

    async def parse(self, record: FileRecord) -> ParsedContent:
        path = Path(record.path)
        ext = path.suffix.lower()

        if record.type == "image":
            return await self.parse_image(path, record.mime_type)
        if ext in TEXT_EXTENSIONS or ext == ".csv":
            return await self.parse_text(path)
        if ext == ".pdf":
            return await self.parse_pdf(path)
        if ext == ".docx":
            return await self.parse_docx(path)
        if ext == ".xlsx":
            return await self.parse_excel(path)

        return ParsedContent(type="text", text=f"[不支持的文件格式: {ext}]")

    async def parse_image(self, path: Path, mime_type: str = "image/png") -> ParsedContent:
        """图片：返回 base64 + OCR 文本"""
        with open(path, "rb") as f:
            image_bytes = f.read()
        image_b64 = base64.b64encode(image_bytes).decode("ascii")

        ocr_text = self._try_ocr(path)

        return ParsedContent(
            type="image",
            image_base64=image_b64,
            mime_type=mime_type,
            ocr_text=ocr_text,
        )

    async def parse_text(self, path: Path) -> ParsedContent:
        """纯文本/代码/CSV：直接读取"""
        try:
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                text = f.read(MAX_TEXT_LENGTH + 1)
            truncated = len(text) > MAX_TEXT_LENGTH
            if truncated:
                text = text[:MAX_TEXT_LENGTH]
            return ParsedContent(type="text", text=text, truncated=truncated)
        except Exception as e:
            return ParsedContent(type="text", text=f"[读取失败: {e}]")

    async def parse_pdf(self, path: Path) -> ParsedContent:
        """PDF 文本提取"""
        try:
            import pypdf
            reader = pypdf.PdfReader(str(path))
            pages = []
            total_len = 0
            for page in reader.pages:
                page_text = page.extract_text() or ""
                if total_len + len(page_text) > MAX_TEXT_LENGTH:
                    pages.append(page_text[:MAX_TEXT_LENGTH - total_len])
                    return ParsedContent(type="text", text="\n".join(pages), truncated=True)
                pages.append(page_text)
                total_len += len(page_text)
            return ParsedContent(type="text", text="\n".join(pages))
        except Exception as e:
            return ParsedContent(type="text", text=f"[PDF 解析失败: {e}]")

    async def parse_docx(self, path: Path) -> ParsedContent:
        """Word 文档文本提取"""
        try:
            import docx
            doc = docx.Document(str(path))
            paragraphs = []
            total_len = 0
            for para in doc.paragraphs:
                text = para.text
                if total_len + len(text) > MAX_TEXT_LENGTH:
                    paragraphs.append(text[:MAX_TEXT_LENGTH - total_len])
                    return ParsedContent(type="text", text="\n".join(paragraphs), truncated=True)
                paragraphs.append(text)
                total_len += len(text)
            return ParsedContent(type="text", text="\n".join(paragraphs))
        except Exception as e:
            return ParsedContent(type="text", text=f"[Word 解析失败: {e}]")

    async def parse_excel(self, path: Path) -> ParsedContent:
        """Excel 转文本表格"""
        try:
            import openpyxl
            wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            lines = []
            total_len = 0
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                lines.append(f"[Sheet: {sheet}]")
                for row in ws.iter_rows(values_only=True):
                    row_text = "\t".join(str(c) if c is not None else "" for c in row)
                    if total_len + len(row_text) > MAX_TEXT_LENGTH:
                        return ParsedContent(type="text", text="\n".join(lines), truncated=True)
                    lines.append(row_text)
                    total_len += len(row_text)
            wb.close()
            return ParsedContent(type="text", text="\n".join(lines))
        except Exception as e:
            return ParsedContent(type="text", text=f"[Excel 解析失败: {e}]")

    def _try_ocr(self, path: Path) -> Optional[str]:
        """尝试 OCR，失败则返回 None"""
        try:
            import pytesseract
            from PIL import Image
            img = Image.open(path)
            text = pytesseract.image_to_string(img, lang="chi_sim+eng")
            if text and text.strip():
                return text.strip()[:MAX_OCR_LENGTH]
            return None
        except Exception:
            return None
