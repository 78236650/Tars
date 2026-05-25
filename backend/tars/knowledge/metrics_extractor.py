"""XLSX/CSV 结构化抽取 — metrics 专路，不让 LLM 抄表。"""
from __future__ import annotations

import csv
import io
import os
from typing import Any, Dict, List, Optional

from .models import DocProfile, GlossaryItem, MetricsTable, ParsedDocument, SectionSummary


def _infer_schema(headers: List[str], rows: List[Dict[str, Any]]) -> Dict[str, str]:
    schema: Dict[str, str] = {}
    for h in headers:
        sample_vals = [r.get(h) for r in rows[:20] if r.get(h) is not None]
        if not sample_vals:
            schema[h] = "text"
            continue
        if all(isinstance(v, (int, float)) and not isinstance(v, bool) for v in sample_vals):
            schema[h] = "number"
        else:
            schema[h] = "text"
    return schema


def parse_csv_tables(file_path: str) -> List[MetricsTable]:
    with open(file_path, "r", encoding="utf-8", errors="replace") as f:
        content = f.read()
    reader = csv.DictReader(io.StringIO(content))
    headers = list(reader.fieldnames or [])
    rows = [dict(row) for row in reader]
    if not headers:
        return []
    return [
        MetricsTable(
            sheet_name=os.path.basename(file_path),
            headers=headers,
            rows=rows,
            inferred_schema=_infer_schema(headers, rows),
        )
    ]


def parse_xlsx_tables(file_path: str) -> List[MetricsTable]:
    try:
        import openpyxl
    except ImportError:
        return []

    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    tables: List[MetricsTable] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_row = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(c).strip() if c is not None else f"col{i}" for i, c in enumerate(header_row)]
        if not any(h for h in headers):
            continue
        data_rows: List[Dict[str, Any]] = []
        for row in rows_iter:
            if not any(cell is not None and str(cell).strip() for cell in row):
                continue
            item = {}
            for i, h in enumerate(headers):
                val = row[i] if i < len(row) else None
                item[h] = val
            data_rows.append(item)
        tables.append(
            MetricsTable(
                sheet_name=sheet_name,
                headers=headers,
                rows=data_rows,
                inferred_schema=_infer_schema(headers, data_rows),
            )
        )
    wb.close()
    return tables


def extract_tables_from_file(file_path: str) -> List[MetricsTable]:
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".csv":
        return parse_csv_tables(file_path)
    if ext == ".xlsx":
        return parse_xlsx_tables(file_path)
    return []


def tables_to_key_facts(tables: List[MetricsTable], *, max_facts: int = 30) -> List[str]:
    facts: List[str] = []
    for table in tables:
        headers = table.headers or []
        for row in table.rows[:50]:
            parts = []
            for h in headers[:6]:
                val = row.get(h)
                if val is None or str(val).strip() == "":
                    continue
                parts.append(f"{h}={val}")
            if parts:
                prefix = f"[{table.sheet_name}] " if table.sheet_name else ""
                facts.append(prefix + " | ".join(parts))
            if len(facts) >= max_facts:
                return facts
    return facts


def build_metrics_profile(
    parsed: ParsedDocument,
    *,
    doc_id: str,
    doc_type: str = "metrics",
    file_name: str = "",
) -> DocProfile:
    """从结构化表格构建 DocProfile，不调用通用 LLM enrich prompt。"""
    tables = list(parsed.metrics_tables or [])
    key_facts = tables_to_key_facts(tables)
    sections = []
    for i, table in enumerate(tables):
        row_count = len(table.rows or [])
        col_names = ", ".join(table.headers[:8])
        sections.append(
            SectionSummary(
                section_id=f"sheet_{i}",
                title=table.sheet_name or f"Sheet {i + 1}",
                summary=f"共 {row_count} 行；列：{col_names}",
                key_facts=tables_to_key_facts([table], max_facts=5),
            )
        )

    one_liner = ""
    if tables:
        one_liner = f"指标表 {len(tables)} 张，共 {sum(len(t.rows) for t in tables)} 行数据"

    return DocProfile(
        doc_id=doc_id,
        doc_type=doc_type,
        title=file_name or doc_id,
        one_liner=one_liner[:160],
        summary=one_liner,
        key_points=[f"工作表：{t.sheet_name}" for t in tables[:10]],
        sections=sections,
        key_facts=key_facts,
        glossary=[],
        qa_pairs=[],
        tags=["metrics", "structured"],
        confidence=0.95 if tables else 0.5,
        parse_warnings=list(parsed.parse_warnings or []),
    )


def enrich_metrics(parsed: ParsedDocument, *, doc_id: str, doc_type: str, file_name: str = "") -> Optional[DocProfile]:
    if not parsed.metrics_tables:
        return None
    return build_metrics_profile(parsed, doc_id=doc_id, doc_type=doc_type, file_name=file_name)
